# -*- coding: utf-8 -*-
"""
Bir rakip bulunamadığında analizin ayakta kalması.

The rival lookup happens at the very end of an analysis that already cost the user
minutes of CPU time. These tests pin down the contract that makes a missing rival a
skipped section rather than a discarded analysis:

  1. get_competitor() always answers with a status; it never raises.
  2. It refuses to hand back an off-niche video just to fill the slot.
  3. save_analysis() still stores the thumbnail/segment/user data that the PDF
     reads back out of the competitor_data column, even with no rival.
"""

import json
import os
import sqlite3
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.services import competitor as comp_mod
from app.services.competitor import (
    COMPETITOR_FOUND_STATUSES,
    COMPETITOR_LOOKUP_FAILED,
    COMPETITOR_MANUAL,
    COMPETITOR_MANUAL_FAILED,
    COMPETITOR_NO_MATCH,
    COMPETITOR_OK,
    CompetitorAnalyzer,
)


# ─── yt-dlp doubles ───────────────────────────────────────────────────────────

class _FakeYDL:
    """Stands in for yt_dlp.YoutubeDL as a context manager."""

    def __init__(self, opts=None):
        self.opts = opts

    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False

    def extract_info(self, url, download=False):  # pragma: no cover - overridden
        raise NotImplementedError


def _install_ydl(monkeypatch, extract_info):
    """
    Point competitor.yt_dlp at a stub whose extract_info is `extract_info`.
    Returns the list of requested URLs so tests can assert how many searches ran.
    """
    calls = []

    class _YDL(_FakeYDL):
        def extract_info(self, url, download=False):
            calls.append(url)
            return extract_info(url)

    monkeypatch.setattr(comp_mod, "YT_DLP_AVAILABLE", True, raising=False)
    monkeypatch.setattr(comp_mod, "yt_dlp", type("m", (), {"YoutubeDL": _YDL}), raising=False)
    return calls


def _install_ydl_sequence(monkeypatch, pages):
    """Answer the Nth search with pages[N], so multi-pass behaviour can be pinned."""
    box = {"n": 0}

    def answer(url):
        i = min(box["n"], len(pages) - 1)
        box["n"] += 1
        return {"entries": pages[i]}

    return _install_ydl(monkeypatch, answer)


def _entry(title, uploader="Baska Kanal", **extra):
    base = {
        "title": title,
        "uploader": uploader,
        "view_count": 12345,
        "like_count": 100,
        "comment_count": 10,
        "tags": [],
        "description": "#etiket bir aciklama",
        "upload_date": "20240101",
    }
    base.update(extra)
    return base


# ─── The shared contract ──────────────────────────────────────────────────────

def _assert_valid_result(result):
    assert isinstance(result, dict), "get_competitor must answer with a dict"
    assert set(result) == {"status", "competitor", "detail"}
    assert isinstance(result["status"], str) and result["status"]
    if result["status"] in COMPETITOR_FOUND_STATUSES:
        assert isinstance(result["competitor"], dict)
    else:
        assert result["competitor"] is None, (
            "a non-found status must not smuggle a competitor through"
        )


def test_found_statuses_are_disjoint_from_failure_statuses():
    failures = {COMPETITOR_NO_MATCH, COMPETITOR_LOOKUP_FAILED, COMPETITOR_MANUAL_FAILED}
    assert set(COMPETITOR_FOUND_STATUSES) == {COMPETITOR_OK, COMPETITOR_MANUAL}
    assert not failures & set(COMPETITOR_FOUND_STATUSES)


def test_missing_yt_dlp_reports_instead_of_raising(monkeypatch):
    monkeypatch.setattr(comp_mod, "YT_DLP_AVAILABLE", False, raising=False)
    result = CompetitorAnalyzer.get_competitor("minecraft", "minecraft, survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_LOOKUP_FAILED


def test_missing_yt_dlp_no_longer_fabricates_a_rival(monkeypatch):
    """The old code answered with an invented 'Sektör Lideri' video."""
    monkeypatch.setattr(comp_mod, "YT_DLP_AVAILABLE", False, raising=False)
    result = CompetitorAnalyzer.get_competitor("minecraft", "minecraft")
    assert result["competitor"] is None


def test_search_crash_reports_lookup_failed(monkeypatch):
    def boom(url):
        raise RuntimeError("network unreachable")

    _install_ydl(monkeypatch, boom)
    result = CompetitorAnalyzer.get_competitor("minecraft", "minecraft, survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_LOOKUP_FAILED
    assert "network unreachable" in result["detail"]


def test_empty_search_results_report_no_match(monkeypatch):
    _install_ydl(monkeypatch, lambda url: {"entries": []})
    result = CompetitorAnalyzer.get_competitor("minecraft", "minecraft, survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_NO_MATCH


def test_off_niche_results_are_refused_rather_than_forced(monkeypatch):
    """
    A trending video from another niche used to be accepted as a last resort,
    which produced nonsense comparisons. It must now be declined.
    """
    _install_ydl(monkeypatch, lambda url: {"entries": [
        _entry("VALORANT radiant rank up"),
        _entry("Iskender kebap tarifi"),
        _entry("Borsa yatirim tavsiyeleri"),
    ]})
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "minecraft, survival", user_title="Minecraft hardcore survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_NO_MATCH


def test_same_niche_result_is_accepted(monkeypatch):
    _install_ydl(monkeypatch, lambda url: {"entries": [
        _entry("Iskender kebap tarifi"),
        _entry("Minecraft hardcore survival ilk gun", uploader="Rakip Kanal", view_count=98765),
    ]})
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "minecraft, survival", user_title="Minecraft hardcore survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_OK
    assert result["competitor"]["views"] == 98765
    assert result["competitor"]["is_fake"] is False


def test_own_channel_is_never_offered_as_a_rival(monkeypatch):
    _install_ydl(monkeypatch, lambda url: {"entries": [
        _entry("Minecraft hardcore survival", uploader="Benim Kanalim"),
    ]})
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "minecraft", channel_name="Benim Kanalim",
        user_title="Minecraft hardcore survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_NO_MATCH


def test_manual_url_bypasses_the_niche_filter(monkeypatch):
    """A link the user typed is a deliberate choice, even across niches."""
    _install_ydl(monkeypatch, lambda url: _entry("Tamamen alakasiz bir video"))
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "minecraft", manual_url="https://www.youtube.com/watch?v=abc123",
        user_title="Minecraft hardcore survival")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_MANUAL
    assert result["competitor"]["is_manual"] is True


def test_unreadable_manual_url_reports_instead_of_raising(monkeypatch):
    def boom(url):
        raise RuntimeError("video is private")

    _install_ydl(monkeypatch, boom)
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "minecraft", manual_url="https://youtu.be/abc123")
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_MANUAL_FAILED


def test_partial_entry_does_not_produce_none_fields(monkeypatch):
    """yt-dlp omits keys freely; the report formats these values directly."""
    _install_ydl(monkeypatch, lambda url: {"title": "Minecraft survival"})
    result = CompetitorAnalyzer.get_competitor(
        "minecraft", "", manual_url="https://youtu.be/abc123")
    c = result["competitor"]
    assert c["views"] == 0 and c["likes"] == 0 and c["comments"] == 0
    assert c["channel"] and c["upload_date"] and isinstance(c["tags"], list)


# ─── Choosing the *best* rival, not the first one ─────────────────────────────

MC_TITLE = "Minecraft hardcore survival"


def _mc(**kw):
    return CompetitorAnalyzer.get_competitor(
        kw.pop("category", "minecraft"), kw.pop("tags", "minecraft, survival"),
        user_title=kw.pop("user_title", MC_TITLE), **kw)


def test_best_candidate_wins_not_the_first_one(monkeypatch):
    """The old code took the first title that shared any word."""
    weak = _entry("Minecraft survival tips", view_count=10_000_000)
    strong = _entry("Minecraft hardcore survival 100 gun", view_count=1_000)
    _install_ydl_sequence(monkeypatch, [[weak, strong]])
    result = _mc()
    assert result["status"] == COMPETITOR_OK
    assert result["competitor"]["title"] == "Minecraft hardcore survival 100 gun", (
        "topic relevance must outrank both result order and view count"
    )


def test_popularity_is_only_a_tie_break(monkeypatch):
    a = _entry("Minecraft hardcore survival serisi", view_count=500)
    b = _entry("Minecraft hardcore survival rehberi", view_count=5_000_000)
    _install_ydl_sequence(monkeypatch, [[a, b]])
    result = _mc()
    assert result["competitor"]["title"] == "Minecraft hardcore survival rehberi"


def test_pick_does_not_depend_on_result_order(monkeypatch):
    entries = [
        _entry("Minecraft survival tips", view_count=900_000),
        _entry("Minecraft hardcore survival 100 gun", view_count=900_000),
        _entry("Minecraft hardcore mode explained", view_count=900_000),
    ]
    _install_ydl_sequence(monkeypatch, [list(entries)])
    first = _mc()["competitor"]["title"]
    _install_ydl_sequence(monkeypatch, [list(reversed(entries))])
    second = _mc()["competitor"]["title"]
    assert first == second, "the same result set must always yield the same rival"


def test_search_queries_are_deterministic():
    def build():
        return comp_mod._MatchContext(
            category="Minecraft oyun", tags="minecraft, survival, hardcore",
            user_title=MC_TITLE).search_queries()

    runs = [build() for _ in range(5)]
    assert all(r == runs[0] for r in runs), (
        "queries came from set iteration before, so they varied between runs"
    )
    assert len(runs[0]) == len(set(runs[0])), "no duplicate searches"
    assert all(q.strip() for q in runs[0])


def test_a_single_short_shared_word_is_not_a_topic_match(monkeypatch):
    _install_ydl_sequence(monkeypatch, [[_entry("Top 100 kek tarifleri")]])
    result = _mc(user_title="Minecraft 100 gun", tags="minecraft")
    assert result["status"] == COMPETITOR_NO_MATCH


def test_rival_the_report_would_warn_about_is_never_picked(monkeypatch):
    """
    compute_kill_switch prints a 'topics do not match' warning in the report. If a
    candidate trips it, picking that candidate would mean warning about our own choice.
    """
    entries = [_entry("Valorant radiant climb", tags=["minecraft"])]
    _install_ydl_sequence(monkeypatch, [entries])
    result = _mc()
    assert result["status"] == COMPETITOR_NO_MATCH
    assert comp_mod.compute_kill_switch(MC_TITLE, "Valorant radiant climb") is True


# ─── Format and language: "comparable" is more than shared words ──────────────

def test_long_video_is_not_offered_as_a_shorts_rival(monkeypatch):
    _install_ydl_sequence(monkeypatch, [[
        _entry("Minecraft hardcore survival full movie", duration=1800),
    ]])
    result = _mc(is_shorts=True, duration_sec=45)
    assert result["status"] == COMPETITOR_NO_MATCH


def test_shorts_analysis_prefers_a_short_rival(monkeypatch):
    long_one = _entry("Minecraft hardcore survival rehberi", duration=150, view_count=9_000_000)
    short_one = _entry("Minecraft hardcore survival anlari", duration=48, view_count=1_000)
    _install_ydl_sequence(monkeypatch, [[long_one, short_one]])
    result = _mc(is_shorts=True, duration_sec=42)
    assert result["competitor"]["title"] == "Minecraft hardcore survival anlari"


def test_long_form_analysis_prefers_a_similar_length_rival(monkeypatch):
    clip = _entry("Minecraft hardcore survival anlari", duration=35, view_count=9_000_000)
    full = _entry("Minecraft hardcore survival rehberi", duration=1100, view_count=1_000)
    _install_ydl_sequence(monkeypatch, [[clip, full]])
    result = _mc(duration_sec=1200)
    assert result["competitor"]["title"] == "Minecraft hardcore survival rehberi"


def test_same_language_rival_is_preferred(monkeypatch):
    """With topic and popularity equal, language decides."""
    tr = _entry("Minecraft hayatta kalma rehberi ikinci gün", view_count=100_000)
    en = _entry("Minecraft hayatta kalma rehberi part two", view_count=100_000)
    _install_ydl_sequence(monkeypatch, [[tr, en]])
    result = _mc(user_title="Minecraft hayatta kalma rehberi ilk gün", tags="minecraft")
    assert result["competitor"]["title"] == "Minecraft hayatta kalma rehberi ikinci gün"


def test_looks_turkish_only_answers_when_it_can_tell():
    assert comp_mod._looks_turkish("Minecraft hayatta kalma ilk gün") is True
    assert comp_mod._looks_turkish("Minecraft survival guide for beginners") is False
    assert comp_mod._looks_turkish("abc") is None


# ─── Spending a second search only when it can help ───────────────────────────

def test_broader_search_runs_only_when_the_narrow_one_finds_nothing(monkeypatch):
    good = _entry("Minecraft hardcore survival 100 gun")
    calls = _install_ydl_sequence(monkeypatch, [[_entry("Iskender kebap tarifi")], [good]])
    result = _mc()
    assert result["status"] == COMPETITOR_OK
    assert result["competitor"]["title"] == good["title"]
    assert len(calls) == 2


def test_a_good_first_match_does_not_trigger_extra_searches(monkeypatch):
    calls = _install_ydl_sequence(monkeypatch, [[_entry("Minecraft hardcore survival 100 gun")]])
    assert _mc()["status"] == COMPETITOR_OK
    assert len(calls) == 1, "a second round trip would cost the user time for nothing"


def test_a_later_search_failing_keeps_the_candidates_already_found(monkeypatch):
    box = {"n": 0}

    def answer(url):
        box["n"] += 1
        if box["n"] == 1:
            return {"entries": [_entry("Iskender kebap tarifi")]}
        raise RuntimeError("network dropped")

    _install_ydl(monkeypatch, answer)
    result = _mc()
    _assert_valid_result(result)
    assert result["status"] == COMPETITOR_NO_MATCH, (
        "we searched successfully once, so this is 'nothing relevant', not 'lookup failed'"
    )


# ─── Turkish text handling ────────────────────────────────────────────────────

def test_turkish_dotted_capital_i_keeps_its_first_letter():
    """
    'İ'.lower() is 'i' plus a combining dot that \\w does not match, so 'İskender'
    used to tokenize as ['i', 'skender'] and silently lose its first letter.
    """
    assert comp_mod.tokenize_keywords("İskender Kebap") == ["iskender", "kebap"]
    assert comp_mod.tokenize_keywords("İSKENDER KEBAP") == ["iskender", "kebap"]
    assert "skender" not in comp_mod.tokenize_keywords("İskender")


def test_rival_is_matched_across_turkish_spelling_variants(monkeypatch):
    _install_ydl_sequence(monkeypatch, [[
        _entry("ISKENDER KEBAP nasil yapilir evde", duration=400),
    ]])
    result = CompetitorAnalyzer.get_competitor(
        "Yemek", "iskender, kebap", user_title="İskender Kebap Nasıl Yapılır",
        duration_sec=420)
    assert result["status"] == COMPETITOR_OK


def test_kill_switch_ignores_diacritics():
    """Two spellings of the same topic are not 'no shared topic'."""
    assert comp_mod.compute_kill_switch("İskender kebap tarifi", "ISKENDER KEBAP") is False
    assert comp_mod.compute_kill_switch("Işık şovu rehberi", "ışık gösterisi") is False
    assert comp_mod.compute_kill_switch("Minecraft survival", "İskender kebap") is True


def test_placeholder_category_does_not_pollute_the_query():
    """A channel with no content type set gets 'Genel', which describes nothing."""
    ctx = comp_mod._MatchContext(category="Genel", tags="minecraft", user_title=MC_TITLE)
    queries = ctx.search_queries()
    assert all("genel" not in q for q in queries)
    assert queries[0] == "minecraft hardcore survival"


def test_the_endpoint_passes_format_and_duration_to_the_lookup():
    src = (ROOT / "server.pyw").read_text(encoding="utf-8", errors="ignore")
    assert 'tech.get("duration"), is_shorts' in src, (
        "without these the scorer cannot judge whether a rival is comparable in length"
    )


def test_extract_core_keywords_still_returns_a_set():
    """server.pyw and check_content_consistency do set algebra on this."""
    kw = comp_mod.extract_core_keywords("Minecraft hardcore survival")
    assert isinstance(kw, set)
    assert kw == {"minecraft", "hardcore", "survival"}
    assert comp_mod.tokenize_keywords("Minecraft hardcore Minecraft survival") == [
        "minecraft", "hardcore", "survival"]


# ─── The persisted data bag ───────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE analyses (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    channel_id INTEGER, video_name TEXT, overall_score REAL, retention_score REAL,
    tech_score REAL, seo_score REAL, thumb_score REAL, peaks INTEGER, viral_score REAL,
    coach_feedback TEXT, competitor_data TEXT, analysis_duration_sec REAL, user_id INTEGER
);
"""


def _save_analysis_blob(result, tmp_path, monkeypatch):
    """Run save_analysis against a throwaway DB and return the stored JSON blob."""
    aiosqlite = pytest.importorskip("aiosqlite")
    from app.services import analysis_engine

    db_path = tmp_path / "t.db"
    with sqlite3.connect(db_path) as raw:
        raw.executescript(SCHEMA)

    async def fake_get_async_db():
        conn = await aiosqlite.connect(db_path)
        return conn

    monkeypatch.setattr(analysis_engine, "get_async_db", fake_get_async_db)

    import asyncio
    analysis_id = asyncio.run(analysis_engine.AnalysisEngine.save_analysis(1, result))

    with sqlite3.connect(db_path) as raw:
        row = raw.execute(
            "SELECT competitor_data FROM analyses WHERE id = ?", (analysis_id,)).fetchone()
    return json.loads(row[0])


BASE_RESULT = {
    "title": "Minecraft hardcore survival",
    "overall_score": 7.5,
    "thumb_data": {"face_detected": True, "faces": [{"dominant_emotion": "happy"}]},
    "viral_segments": [{"start": 10, "end": 20}],
    "user_meta": {
        "user_title_len": 27,
        "user_tags": ["minecraft", "survival"],
        "user_hashtags": ["minecraft"],
        "user_description": "#minecraft aciklama",
    },
}


def test_report_data_survives_a_missing_rival(tmp_path, monkeypatch):
    """
    competitor_data doubles as the report's data bag: the PDF reads the thumbnail
    analysis, viral segments and the user's own tags back out of it. None of that
    may be lost just because no rival was found.
    """
    result = dict(BASE_RESULT, competitor_data=None, competitor_status=COMPETITOR_NO_MATCH)
    blob = _save_analysis_blob(result, tmp_path, monkeypatch)

    assert blob["has_competitor"] is False
    assert blob["competitor_status"] == COMPETITOR_NO_MATCH
    assert blob["_thumb_data"]["faces"], "thumbnail analysis must still be stored"
    assert blob["_viral_segments"], "viral segments must still be stored"
    assert blob["user_tags"] == ["minecraft", "survival"]
    assert blob["user_description"] == "#minecraft aciklama"
    assert blob["face_detected"] is True
    assert "title" not in blob, "no rival means no rival title to render"


def test_rival_is_stored_alongside_the_report_data(tmp_path, monkeypatch):
    result = dict(
        BASE_RESULT,
        competitor_data={"title": "Rakip video", "channel": "Rakip Kanal", "views": 500},
        competitor_status=COMPETITOR_OK,
    )
    blob = _save_analysis_blob(result, tmp_path, monkeypatch)

    assert blob["has_competitor"] is True
    assert blob["title"] == "Rakip video"
    assert blob["_thumb_data"]["faces"]
    assert blob["user_tags"] == ["minecraft", "survival"]


def test_save_analysis_does_not_mutate_the_api_payload(tmp_path, monkeypatch):
    """The response goes to the browser; internal '_' keys shouldn't ride along."""
    payload = {"title": "Rakip video", "channel": "Rakip Kanal", "views": 500}
    result = dict(BASE_RESULT, competitor_data=payload, competitor_status=COMPETITOR_OK)
    _save_analysis_blob(result, tmp_path, monkeypatch)
    assert "_thumb_data" not in payload
    assert "has_competitor" not in payload


# ─── Wiring ───────────────────────────────────────────────────────────────────

def test_analyze_endpoint_treats_a_missing_rival_as_non_fatal():
    src = (ROOT / "server.pyw").read_text(encoding="utf-8", errors="ignore")
    assert "COMPETITOR_FOUND_STATUSES" in src, "the endpoint must branch on the status"
    assert '"competitor_status": competitor_status' in src, "the UI needs the status"
    assert "Henüz kıyaslanacak başka bir rakip" not in src, "the fatal error must be gone"


def test_pdf_skips_only_the_head_to_head_section():
    src = (ROOT / "server.pyw").read_text(encoding="utf-8", errors="ignore")
    assert "NO_COMPETITOR_SKIP" in src
    # The SEO check-up is about the user's own video, so it has to be appended
    # before the rival guard bails out.
    assert src.index("elements.append(Paragraph(checkup_txt, normal_s))") < \
        src.index('raise ValueError("NO_COMPETITOR_SKIP")')


def test_ui_explains_a_missing_rival_for_every_status():
    from openpyxl import load_workbook

    ws = load_workbook(ROOT / "translations.xlsx", data_only=True)["ui"]
    header = [c.value for c in ws[1]]
    keys = {str(ws.cell(row=r, column=header.index("key") + 1).value or "").strip(): r
            for r in range(2, ws.max_row + 1)}

    app_js = (ROOT / "static" / "App.js").read_text(encoding="utf-8", errors="ignore")
    for key in ("compSkipped", "compNoMatch", "compUrlFailed", "compLookupFailed"):
        assert key in keys, f"translations.xlsx/ui is missing {key}"
        assert f"'{key}'" in app_js, f"App.js never renders {key}"
        for lang in ("tr", "en", "es"):
            cell = ws.cell(row=keys[key], column=header.index(lang) + 1).value
            assert str(cell or "").strip(), f"{key} has no {lang} text"


# ─── Reopening a saved analysis ───────────────────────────────────────────────

APP_JS = ROOT / "static" / "App.js"


def _eval_js_expression(expr: str, cases: list):
    """
    Run an expression lifted straight out of App.js against sample blobs.

    Reading the expression from the file instead of restating it here means the
    test cannot quietly drift away from the shipped code.
    """
    import subprocess
    import tempfile

    script = (
        "function decide(blob) { return %s; }\n"
        "const cases = %s;\n"
        "console.log(JSON.stringify(cases.map(c => decide(c))));\n"
    ) % (expr, json.dumps(cases))

    with tempfile.NamedTemporaryFile("w", suffix=".js", delete=False, encoding="utf-8") as fh:
        fh.write(script)
        path = fh.name
    try:
        out = subprocess.run(["node", path], capture_output=True, text=True, timeout=60)
    finally:
        os.unlink(path)
    assert out.returncode == 0, f"node failed: {out.stderr}"
    return json.loads(out.stdout.strip())


def _extract(pattern: str) -> str:
    import re

    src = APP_JS.read_text(encoding="utf-8", errors="ignore")
    m = re.search(pattern, src)
    assert m, f"App.js no longer contains {pattern!r}"
    return m.group(1)


# A rival-less row still carries the thumbnail analysis and the user's own tags,
# so the blob is truthy and cannot be used as a rival on its own.
NO_RIVAL_BLOB = {"has_competitor": False, "competitor_status": "no_confident_match",
                 "_thumb_data": {"faces": []}, "user_tags": ["minecraft"]}
RIVAL_BLOB = {"has_competitor": True, "competitor_status": "ok",
              "title": "Rakip video", "channel": "Rakip Kanal", "views": 500}
LEGACY_BLOB = {"title": "Eski rakip", "channel": "Eski Kanal", "views": 900}


def test_saved_analysis_does_not_mistake_the_data_bag_for_a_rival():
    """Without this the history view throws on comp.channel and renders nothing."""
    expr = _extract(r"const hasComp = ([^;]+);")
    results = _eval_js_expression(expr, [
        NO_RIVAL_BLOB, RIVAL_BLOB, LEGACY_BLOB, None,
        {"has_competitor": True},  # flagged but unusable
    ])
    assert results == [False, True, True, False, False]


def test_reopened_analysis_keeps_the_reason_there_is_no_rival():
    expr = _extract(r"competitor_status:\s*(.+?),\n")
    results = _eval_js_expression(
        expr.replace("hasComp", "(blob && blob.has_competitor !== false && blob.title)"),
        [NO_RIVAL_BLOB, RIVAL_BLOB, LEGACY_BLOB, None])
    assert results[0] == "no_confident_match", "the note must explain which case this is"
    assert results[1] == "ok"
    assert results[2] == "ok", "a legacy row with a rival is simply 'ok'"
    assert results[3] == "no_confident_match"


def test_thumbnail_and_segments_survive_reopening_without_a_rival():
    """These live in the blob, so they must be read from it even with no rival."""
    src = APP_JS.read_text(encoding="utf-8", errors="ignore")
    assert "(blob && blob._viral_segments) || []" in src
    assert "(blob && blob._thumb_data) || {}" in src


def test_vs_card_tolerates_a_rival_with_missing_fields():
    src = APP_JS.read_text(encoding="utf-8", errors="ignore")
    assert "(comp.channel || '').toUpperCase()" in src
    assert "(comp.views || 0).toLocaleString()" in src


def test_the_fabricated_rival_warning_is_gone():
    """Nothing invents a rival any more, so the warning had no way to fire."""
    src = APP_JS.read_text(encoding="utf-8", errors="ignore")
    assert "fakeWarningHTML" not in src


def test_view_count_label_is_translated():
    from openpyxl import load_workbook

    src = APP_JS.read_text(encoding="utf-8", errors="ignore")
    assert "t('compViews')" in src
    assert "} İzlenme" not in src, "the label used to be hardcoded Turkish"

    ws = load_workbook(ROOT / "translations.xlsx", data_only=True)["ui"]
    header = [c.value for c in ws[1]]
    rows = {str(ws.cell(row=r, column=header.index("key") + 1).value or "").strip(): r
            for r in range(2, ws.max_row + 1)}
    assert "compViews" in rows
    for lang in ("tr", "en", "es"):
        assert str(ws.cell(row=rows["compViews"], column=header.index(lang) + 1).value or "").strip()


def test_pdf_explains_a_missing_rival():
    from openpyxl import load_workbook

    ws = load_workbook(ROOT / "translations.xlsx", data_only=True)["pdf"]
    header = [c.value for c in ws[1]]
    found = {}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=header.index("key") + 1).value or "").strip() == "no_competitor_desc":
            found = {lang: ws.cell(row=r, column=header.index(lang) + 1).value
                     for lang in ("tr", "en", "es")}
            break
    assert found, "translations.xlsx/pdf is missing no_competitor_desc"
    for lang, text in found.items():
        assert str(text or "").strip(), f"no_competitor_desc has no {lang} text"
